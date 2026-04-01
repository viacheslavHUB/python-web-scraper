import csv
import json
import logging
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_URL = "https://books.toscrape.com/"
START_URL = urljoin(BASE_URL, "catalogue/page-1.html")

OUTPUT_DIR = Path("output")
CSV_FILE = OUTPUT_DIR / "books.csv"
JSON_FILE = OUTPUT_DIR / "books.json"
XLSX_FILE = OUTPUT_DIR / "books.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0"
}

RATING_MAP = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5,
}

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s: %(message)s"
)


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)

    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset({"GET"}),
    )

    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    return session


def get_soup(session: requests.Session, url: str) -> BeautifulSoup:
    response = session.get(url, timeout=15)
    response.raise_for_status()
    return BeautifulSoup(response.text, "html.parser")


def parse_rating(classes: list[str]) -> int:
    for class_name in classes:
        if class_name in RATING_MAP:
            return RATING_MAP[class_name]
    return 0


def parse_book(article: Tag, page_url: str) -> dict[str, str | int]:
    h3_tag = article.find("h3")
    title_tag = h3_tag.find("a") if isinstance(h3_tag, Tag) else None
    price_tag = article.find("p", class_="price_color")
    stock_tag = article.find("p", class_="instock availability")
    rating_tag = article.find("p", class_="star-rating")

    if not all(isinstance(tag, Tag) for tag in [title_tag, price_tag, stock_tag, rating_tag]):
        raise ValueError("Book card has unexpected structure")

    relative_link = title_tag["href"]
    product_url = urljoin(page_url, relative_link)

    return {
        "title": title_tag.get("title", "").strip(),
        "price": price_tag.get_text(strip=True),
        "availability": stock_tag.get_text(" ", strip=True),
        "rating": parse_rating(rating_tag.get("class", [])),
        "product_url": product_url,
    }


def scrape_page(session: requests.Session, url: str) -> tuple[list[dict[str, str | int]], str | None]:
    soup = get_soup(session, url)

    books: list[dict[str, str | int]] = []
    articles = soup.find_all("article", class_="product_pod")

    for article in articles:
        if isinstance(article, Tag):
            books.append(parse_book(article, url))

    next_button = soup.find("li", class_="next")
    next_page_url = None

    if isinstance(next_button, Tag):
        next_link = next_button.find("a")
        if isinstance(next_link, Tag):
            next_page_url = urljoin(url, next_link.get("href", ""))

    return books, next_page_url


def scrape_catalog(session: requests.Session, start_url: str) -> list[dict[str, str | int]]:
    all_books: list[dict[str, str | int]] = []
    current_url = start_url
    page_number = 1

    while current_url:
        logging.info(f"Scraping page {page_number}: {current_url}")

        try:
            books, next_page_url = scrape_page(session, current_url)
            all_books.extend(books)
            current_url = next_page_url
            page_number += 1
        except requests.RequestException as error:
            logging.error(f"Request failed: {error}")
            break
        except Exception as error:
            logging.error(f"Unexpected error: {error}")
            break

    return all_books


def save_to_csv(data: list[dict[str, str | int]], file_path: Path) -> None:
    if not data:
        logging.warning("No data to save to CSV.")
        return

    with file_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    logging.info(f"CSV saved: {file_path}")


def save_to_json(data: list[dict[str, str | int]], file_path: Path) -> None:
    if not data:
        logging.warning("No data to save to JSON.")
        return

    with file_path.open("w", encoding="utf-8") as file:
        json.dump(data, file, indent=2, ensure_ascii=False)

    logging.info(f"JSON saved: {file_path}")


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def save_to_excel(data: list[dict[str, str | int]], file_path: Path) -> None:
    if not data:
        logging.warning("No data to save to Excel.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Books"

    headers = list(data[0].keys())
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for row in data:
        values = [row[key] for key in headers]
        sheet.append(values)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    product_url_col = headers.index("product_url") + 1
    for row_idx in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=product_url_col)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    autosize_columns(sheet)

    workbook.save(file_path)
    logging.info(f"Excel saved: {file_path}")


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)

    logging.info("Starting scraper...")

    with create_session() as session:
        books = scrape_catalog(session, START_URL)

    logging.info(f"Total books scraped: {len(books)}")

    save_to_csv(books, CSV_FILE)
    save_to_json(books, JSON_FILE)
    save_to_excel(books, XLSX_FILE)

    logging.info("Done.")


if __name__ == "__main__":
    main() 